from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from ..database import get_db
from ..models.user import User
from ..schemas.patient import PatientCreate, PatientUpdate, PatientInDB
from ..crud import patient as crud_patient
from ..routers.auth import get_current_user
from ..services import geocoding_service
import csv
import io
import codecs
from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
from pydantic import BaseModel

router = APIRouter(prefix="/patients", tags=["patients"])

@router.get("/recurring", response_model=List[PatientInDB])
def get_recurring_patients(
    date_str: str, # Format YYYY-MM-DD
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from datetime import datetime
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    # Python weekday(): 0=Mon, 1=Tue, ..., 6=Sun
    # Let's use 1=Mon, 2=Tue, 3=Wed, 4=Thu, 5=Fri, 6=Sat, 7=Sun
    day_mapping = {0: "1", 1: "2", 2: "3", 3: "4", 4: "5", 5: "6", 6: "7"}
    day_str = day_mapping[dt.weekday()]
    
    from sqlalchemy.orm import joinedload
    patients = db.query(crud_patient.model).options(joinedload(crud_patient.model.skills_required)).filter(
        crud_patient.model.is_active == True,
        crud_patient.model.recurrence_days.like(f"%{day_str}%")
    ).all()
    
    return patients

@router.get("/", response_model=List[PatientInDB])
def read_patients(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    filter_active: bool = True, # default show active only
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # TODO: Implement advanced filtering for 'is_active' and other fields within CRUD
    # Custom query because crud is generic
    from sqlalchemy.orm import joinedload
    query = db.query(crud_patient.model).options(joinedload(crud_patient.model.skills_required))
    
    if filter_active:
        query = query.filter(crud_patient.model.is_active == True)
        
    if search:
        search_term = f"%{search}%"
        from sqlalchemy import or_
        query = query.filter(
            or_(
                crud_patient.model.nom.ilike(search_term),
                crud_patient.model.prenom.ilike(search_term),
                crud_patient.model.adresse.ilike(search_term)
            )
        )
        
    count = query.count()
    patients = query.offset(skip).limit(limit).all()
    
    # Add count header? Or just return list
    return patients

@router.post("/", response_model=PatientInDB)
async def create_patient(
    patient_in: PatientCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Auto-geocode on create if needed
    if patient_in.latitude and patient_in.longitude:
        coords = (patient_in.latitude, patient_in.longitude)
    else:
        coords = await geocoding_service.geocode(patient_in.adresse)
    
    patient = crud_patient.create(db, obj_in=patient_in)
    
    if coords and not (patient_in.latitude and patient_in.longitude):
        patient.latitude = coords[0]
        patient.longitude = coords[1]
        db.commit()
        db.refresh(patient)
        
    return patient

@router.get("/{id}", response_model=PatientInDB)
def read_patient(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    patient = crud_patient.get(db, id=id)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient

@router.put("/{id}", response_model=PatientInDB)
async def update_patient(
    id: int,
    patient_in: PatientUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    patient = crud_patient.get(db, id=id)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    old_address = patient.adresse
    updated_patient = crud_patient.update(db, db_obj=patient, obj_in=patient_in)
    
    # If coordinates are manually updated in payload, we don't need to geocode
    if patient_in.latitude and patient_in.longitude:
        pass
    elif patient_in.adresse and patient_in.adresse != old_address:
         coords = await geocoding_service.geocode(patient_in.adresse)
         if coords:
            updated_patient.latitude = coords[0]
            updated_patient.longitude = coords[1]
            db.commit()
            db.refresh(updated_patient)

    return updated_patient

@router.delete("/{id}", response_model=PatientInDB)
def delete_patient(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    patient = crud_patient.get(db, id=id)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    # Soft delete
    patient.is_active = False
    db.commit()
    db.refresh(patient)
    return patient

@router.get("/{id}/history")
def get_patient_history(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from ..models.tournee import Tournee, TourneeDetails
    
    # Get all stops for this patient
    history = db.query(TourneeDetails).join(Tournee).filter(
        TourneeDetails.patient_id == id
    ).order_by(Tournee.date.desc()).all()
    
    res = []
    for h in history:
        res.append({
            "date": h.tournee.date,
            "heure_arrivee": h.heure_arrivee_estimee,
            "statut": h.statut,
            "nurse": h.tournee.user.username if h.tournee.user else "Non assigné",
            "notes": h.notes
        })
    return res

@router.post("/deactivate-all")
def deactivate_all_patients(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Soft delete all active patients
    db.query(crud_patient.model).filter(crud_patient.model.is_active == True).update({"is_active": False})
    db.commit()
    return {"message": "Tous les patients ont été désactivés"}

@router.get("/import/template")
def get_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modele_Import"
    headers = ["nom", "prenom", "adresse", "telephone", "email", "date_naissance", "a_jeun"]
    ws.append(headers)
    ws.append(["Dupont", "Jean", "10 Rue de Paris, 75001 Paris", "0601020304", "jean.dupont@mail.com", "15/05/1980", "Oui"])
    ws.append(["Martin", "Sophie", "5 Avenue des Champs-Elysées, 75008 Paris", "0611223344", "sophie.m@mail.com", "22/10/1992", "Non"])
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "Modele_Import_Patients.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/import")
async def import_patients(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from ..models.setting import AppSetting
    import pandas as pd
    
    settings = db.query(AppSetting).first()
    import_mode = settings.import_mode if settings else "skip"
    
    content = await file.read()
    filename = file.filename.lower()
    
    stats = {"added": 0, "updated": 0, "errors": 0, "deleted": 0}
    
    if import_mode == "replace":
        # Desactivate all active patients before import
        stats['deleted'] = db.query(crud_patient.model).filter(crud_patient.model.is_active == True).update({"is_active": False})
        db.commit()

    rows = []
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            df = pd.read_excel(io.BytesIO(content))
            df = df.where(pd.notnull(df), None) # Replace NaN with None
            rows = df.to_dict('records')
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erreur de lecture du fichier Excel: {str(e)}")
    else:
        # CSV Handling
        if content.startswith(codecs.BOM_UTF8):
            content = content[len(codecs.BOM_UTF8):]
        stream = io.StringIO(content.decode("utf-8"))
        sample = stream.read(1024)
        stream.seek(0)
        delimiter = ';' if ';' in sample else ','
        csv_reader = csv.DictReader(stream, delimiter=delimiter)
        rows = list(csv_reader)

    from datetime import datetime
    for row in rows:
        try:
            def get_val(keys):
                for k in keys:
                    if k in row and row[k] is not None and str(row[k]).strip() != "":
                        return str(row[k]).strip()
                return None

            nom = get_val(['nom', 'Nom', 'lastname'])
            prenom = get_val(['prenom', 'Prénom', 'firstname'])
            adresse = get_val(['adresse', 'Adresse', 'address'])
            
            if not nom or not prenom or not adresse:
                stats['errors'] += 1
                continue
                
            # Date parsing
            date_naiss = None
            dob_str = get_val(['date_naissance', 'ddn', 'dob', 'Date de Naissance'])
            if dob_str:
                try:
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"]:
                        try:
                            date_naiss = datetime.strptime(dob_str, fmt).date()
                            break
                        except: pass
                except: pass

            patient_in = PatientCreate(
                nom=nom,
                prenom=prenom,
                adresse=adresse,
                telephone=get_val(['telephone', 'tel', 'phone', 'Téléphone']),
                email=get_val(['email', 'mail', 'Email']),
                date_naissance=date_naiss,
                test_a_jeun=str(get_val(['a_jeun', 'A jeun', 'test_a_jeun'])).lower() in ['true', '1', 'oui', 'yes'],
                is_active=True
            )
            
            existing = db.query(crud_patient.model).filter(
                crud_patient.model.nom == patient_in.nom,
                crud_patient.model.prenom == patient_in.prenom,
                crud_patient.model.is_active == True
            ).first()
            
            if not existing:
                crud_patient.create(db, obj_in=patient_in)
                stats['added'] += 1
            else:
                if import_mode == "update":
                    update_data = patient_in.dict(exclude_unset=True)
                    crud_patient.update(db, db_obj=existing, obj_in=update_data)
                    stats['updated'] += 1
                else:
                    # skip
                    pass
        except Exception as e:
            stats['errors'] += 1
            
    return stats

class PurgeRequest(BaseModel):
    purge_history: bool = False
    purge_inactive_patients: bool = False

@router.get("/export/excel")
def export_patients_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Non autorisé.")
        
    patients = db.query(crud_patient.model).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients_Export"
    
    headers = ["ID", "Nom", "Prénom", "Adresse", "Téléphone", "Email", "Date de Naissance", "Actif", "Latitude", "Longitude", "À jeun"]
    ws.append(headers)
    
    for p in patients:
        ws.append([
            p.id, p.nom, p.prenom, p.adresse, p.telephone, p.email, 
            p.date_naissance.strftime("%Y-%m-%d") if p.date_naissance else "",
            "Oui" if p.is_active else "Non",
            p.latitude, p.longitude, 
            "Oui" if p.test_a_jeun else "Non"
        ])
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/purge")
def purge_data(
    req: PurgeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="Non autorisé. Réservé aux administrateurs.")
        
    from ..models.setting import AppSetting
    from ..models.tournee import Tournee, TourneeDetails
    
    settings = db.query(AppSetting).first()
    days = settings.purge_after_days if settings else 30
    cutoff_date = datetime.now().date() - timedelta(days=days)
    
    deleted_history = 0
    deleted_patients = 0
    
    if req.purge_history:
        old_tournees = db.query(Tournee).filter(Tournee.date < cutoff_date).all()
        for t in old_tournees:
            db.query(TourneeDetails).filter(TourneeDetails.tournee_id == t.id).delete()
            db.delete(t)
            deleted_history += 1
            
    if req.purge_inactive_patients:
        deleted_patients = db.query(crud_patient.model).filter(crud_patient.model.is_active == False).delete()
        
    db.commit()
    
    msg = []
    if req.purge_history: msg.append(f"{deleted_history} anciennes tournées supprimées.")
    if req.purge_inactive_patients: msg.append(f"{deleted_patients} patients inactifs supprimés.")
    
    if not msg:
        msg.append("Aucune action de purge sélectionnée.")
        
    return {"message": " ".join(msg)}
